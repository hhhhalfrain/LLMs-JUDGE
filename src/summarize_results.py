from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# ✅ 硬编码配置区（按需改这里）
# ============================================================

# 建议把这个脚本放在你的仓库里，比如：<PROJECT_ROOT>/tools/export_discussion_scores_to_xlsx.py
# PROJECT_ROOT 会自动取该脚本上两级目录（与 batch_run.py 一致）。
PROJECT_ROOT = Path(__file__).resolve().parents[1]

# 若你只想导出某个 batch：填写 batch_id；若想扫所有 runs：设为 None
BATCH_ID: Optional[str] = "4agents exp4"

# 扫描范围
if BATCH_ID:
    SEARCH_ROOT = PROJECT_ROOT / "runs" / "batch" / BATCH_ID
else:
    SEARCH_ROOT = PROJECT_ROOT / "runs"

# result.json 的相对位置一般是 outputs/<run_id>/result.json
RESULT_GLOB = "**/result.json"

# 输出路径
OUT_XLSX = PROJECT_ROOT / "runs" / "exports" / (f"export_{BATCH_ID}.xlsx" if BATCH_ID else "export_all.xlsx")


# ============================================================
# 工具函数
# ============================================================

def _read_json(p: Path) -> Dict[str, Any]:
    with p.open("r", encoding="utf-8") as f:
        return json.load(f)


def _get(d: Dict[str, Any], path: str, default: Any = None) -> Any:
    cur: Any = d
    for k in path.split("."):
        if not isinstance(cur, dict) or k not in cur:
            return default
        cur = cur[k]
    return cur


def _to_bool(x: Any, default: bool = False) -> bool:
    if x is None:
        return default
    if isinstance(x, bool):
        return x
    if isinstance(x, (int, float)):
        return bool(int(x))
    s = str(x).strip().lower()
    if s in ("1", "true", "yes", "y", "t"):
        return True
    if s in ("0", "false", "no", "n", "f"):
        return False
    return default


def _to_int(x: Any, default: int = 0) -> int:
    try:
        return int(x)
    except Exception:
        return default


def _to_float(x: Any, default: Optional[float] = None) -> Optional[float]:
    try:
        if x is None:
            return default
        return float(x)
    except Exception:
        return default


def _safe_str(x: Any, default: str = "") -> str:
    if x is None:
        return default
    return str(x)


def _infer_batch_id(path: Path) -> str:
    """
    从路径推断 batch_id（如果是在 runs/batch/<BATCH_ID>/... 下）。
    """
    parts = [p for p in path.parts]
    # .../runs/batch/<BATCH_ID>/...
    for i in range(len(parts) - 2):
        if parts[i] == "runs" and parts[i + 1] == "batch":
            return parts[i + 2]
    return ""


def _extract_run_id(result_path: Path) -> str:
    # 一般：.../outputs/<run_id>/result.json
    return result_path.parent.name


def _extract_book_name(obj: Dict[str, Any]) -> str:
    # 兼容：book_name 或 metadata.book_name 或 metadata.title
    return (
        _safe_str(obj.get("book_name"))
        or _safe_str(_get(obj, "metadata.book_name"))
        or _safe_str(_get(obj, "metadata.title"))
        or "UNKNOWN"
    )


def _extract_config(obj: Dict[str, Any]) -> Dict[str, Any]:
    # 旧格式：config = _safe_cfg_dump(cfg)
    cfg = obj.get("config", {}) or {}
    if not isinstance(cfg, dict):
        return {}
    return cfg


def _extract_experiment_cfg(cfg: Dict[str, Any]) -> Dict[str, Any]:
    exp = cfg.get("experiment", {}) or {}
    if isinstance(exp, dict) and exp:
        return exp
    # 兜底：有些版本可能把 experiment 平铺在 config 里
    return cfg


def _discussion_scores_from_agent(agent: Dict[str, Any]) -> Tuple[Optional[float], Dict[int, float]]:
    """
    返回：
    - pre_score (round 0)
    - round_score_map: {round_index: score_after_round}
    """
    pre = _to_float(agent.get("pre_discussion_score"), default=None)
    round_map: Dict[int, float] = {}

    disc = agent.get("discussion", []) or []
    if isinstance(disc, list):
        for ev in disc:
            if not isinstance(ev, dict):
                continue
            r = _to_int(ev.get("round", ev.get("r", None)), default=-1)
            sc = _to_float(ev.get("score", None), default=None)
            if r >= 1 and sc is not None:
                round_map[r] = float(sc)

    return pre, round_map


def _format_sheet(ws, freeze_row: int = 1) -> None:
    # 冻结窗格
    ws.freeze_panes = ws[f"A{freeze_row+1}"]

    # 标题行样式
    header_fill = PatternFill("solid", fgColor="1F4E79")  # 深蓝
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ws.row_dimensions[1].height = 22

    # 自动筛选
    ws.auto_filter.ref = ws.dimensions

    # 自动列宽（限制最大宽度，避免超级长文本撑爆）
    max_width = 60
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        best = 10
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            s = str(v)
            best = max(best, min(max_width, len(s) + 2))
        ws.column_dimensions[col_letter].width = best


def _apply_number_formats(ws, score_cols: List[str]) -> None:
    """
    给分数列统一设置 0.0 的显示格式
    score_cols: 列名列表（会在第一行里匹配）
    """
    header = [c.value for c in ws[1]]
    col_idx = {name: i + 1 for i, name in enumerate(header) if name is not None}

    for name in score_cols:
        if name not in col_idx:
            continue
        c = col_idx[name]
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0"


# ============================================================
# 主逻辑
# ============================================================

def main() -> None:
    if not SEARCH_ROOT.exists():
        raise RuntimeError(f"SEARCH_ROOT 不存在：{SEARCH_ROOT}")

    result_paths = sorted(SEARCH_ROOT.glob(RESULT_GLOB))
    if not result_paths:
        raise RuntimeError(f"在 {SEARCH_ROOT} 下未找到 result.json（glob={RESULT_GLOB}）")

    experiments_rows: List[Dict[str, Any]] = []
    scores_long_rows: List[Dict[str, Any]] = []
    errors_rows: List[Dict[str, Any]] = []

    max_round_seen = 0

    for rp in result_paths:
        try:
            obj = _read_json(rp)
            run_id = _extract_run_id(rp)
            book_name = _extract_book_name(obj)
            batch_id = _infer_batch_id(rp)

            cfg = _extract_config(obj)
            exp = _extract_experiment_cfg(cfg)

            method = _safe_str(exp.get("method", ""))
            use_persona = _to_bool(exp.get("use_persona", None), default=False)
            use_discussion = _to_bool(exp.get("use_discussion", None), default=False)
            use_interest_filter = _to_bool(exp.get("use_interest_filter", None), default=False)

            discussion_rounds = _to_int(exp.get("discussion_rounds", 0), default=0)
            discussion_window = _to_int(exp.get("discussion_window", 0), default=0)
            n_agents_cfg = _to_int(exp.get("n_agents", None), default=0)
            score_decimals = _to_int(exp.get("score_decimals", None), default=0)
            discussion_affects_score = _to_bool(exp.get("discussion_affects_score", None), default=False)
            chapter_batch_size = _to_int(exp.get("chapter_batch_size", None), default=0)

            base_url = _safe_str(_get(cfg, "llm.base_url", ""))
            run_ts = _safe_str(obj.get("run_ts", ""))

            agents = obj.get("agents", []) or []
            if not isinstance(agents, list):
                agents = []

            n_agents_total = _to_int(_get(obj, "aggregate.n_agents_total", None), default=len(agents) or n_agents_cfg)

            experiments_rows.append(
                {
                    "batch_id": batch_id,
                    "run_id": run_id,
                    "book_name": book_name,
                    "method": method,
                    "use_persona": use_persona,
                    "use_discussion": use_discussion,
                    "use_interest_filter": use_interest_filter,
                    "discussion_rounds_cfg": discussion_rounds,
                    "discussion_window": discussion_window,
                    "n_agents_total": n_agents_total,
                    "score_decimals": score_decimals,
                    "discussion_affects_score": discussion_affects_score,
                    "chapter_batch_size": chapter_batch_size,
                    "base_url": base_url,
                    "run_ts": run_ts,
                    "result_path": str(rp.relative_to(PROJECT_ROOT)),
                }
            )

            # scores_long：round=0 记录 pre_discussion_score
            for agent in agents:
                if not isinstance(agent, dict):
                    continue
                uid = _safe_str(agent.get("agent_uuid", agent.get("uuid", "")))
                kept = _to_bool(agent.get("kept", True), default=True)

                pre_score, round_map = _discussion_scores_from_agent(agent)
                # round=0
                scores_long_rows.append(
                    {
                        "batch_id": batch_id,
                        "run_id": run_id,
                        "book_name": book_name,
                        "agent_uuid": uid,
                        "kept": kept,
                        "round": 0,
                        "score": pre_score,
                        "username": None,
                        "message": None,
                    }
                )

                # rounds >= 1
                disc = agent.get("discussion", []) or []
                # 为了把 username/message 一起带出来：先建 round -> (username,message,score)
                round_details: Dict[int, Dict[str, Any]] = {}
                if isinstance(disc, list):
                    for ev in disc:
                        if not isinstance(ev, dict):
                            continue
                        r = _to_int(ev.get("round", None), default=-1)
                        if r < 1:
                            continue
                        round_details[r] = {
                            "username": _safe_str(ev.get("username", ""), default="") or None,
                            "message": _safe_str(ev.get("message", ""), default="") or None,
                            "score": _to_float(ev.get("score", None), default=None),
                        }

                for r, sc in sorted(round_map.items(), key=lambda x: x[0]):
                    max_round_seen = max(max_round_seen, r)
                    det = round_details.get(r, {})
                    scores_long_rows.append(
                        {
                            "batch_id": batch_id,
                            "run_id": run_id,
                            "book_name": book_name,
                            "agent_uuid": uid,
                            "kept": kept,
                            "round": r,
                            "score": sc,
                            "username": det.get("username", None),
                            "message": det.get("message", None),
                        }
                    )

        except Exception as e:
            errors_rows.append(
                {
                    "result_path": str(rp),
                    "error_type": type(e).__name__,
                    "error": str(e),
                }
            )

    # 构建 DataFrame
    experiments_df = pd.DataFrame(experiments_rows)
    scores_long_df = pd.DataFrame(scores_long_rows)
    errors_df = pd.DataFrame(errors_rows)

    # scores_wide：透视
    # 统一列名 score_r0..score_rN（N=全局最大讨论轮）
    if not scores_long_df.empty:
        pivot = scores_long_df.pivot_table(
            index=["batch_id", "run_id", "book_name", "agent_uuid", "kept"],
            columns="round",
            values="score",
            aggfunc="last",
        ).reset_index()

        # 重命名 round 列
        new_cols = []
        for c in pivot.columns:
            if isinstance(c, (int, float)):
                new_cols.append(f"score_r{int(c)}")
            else:
                new_cols.append(str(c))
        pivot.columns = new_cols

        # 确保 score_r0..score_rN 都存在
        for r in range(0, max_round_seen + 1):
            col = f"score_r{r}"
            if col not in pivot.columns:
                pivot[col] = None

        # 调整列顺序：固定 key + score_r0..
        key_cols = ["batch_id", "run_id", "book_name", "agent_uuid", "kept"]
        score_cols = [f"score_r{r}" for r in range(0, max_round_seen + 1)]
        pivot = pivot[key_cols + score_cols]

        scores_wide_df = pivot
    else:
        scores_wide_df = pd.DataFrame(columns=["batch_id", "run_id", "book_name", "agent_uuid", "kept"])

    # 写入 xlsx
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        experiments_df.to_excel(writer, sheet_name="experiments", index=False)
        scores_long_df.to_excel(writer, sheet_name="scores_long", index=False)
        scores_wide_df.to_excel(writer, sheet_name="scores_wide", index=False)
        errors_df.to_excel(writer, sheet_name="errors", index=False)

    # 二次格式化（openpyxl）
    wb = load_workbook(OUT_XLSX)

    # experiments
    ws = wb["experiments"]
    _format_sheet(ws)
    _apply_number_formats(
        ws,
        score_cols=[
            "discussion_rounds_cfg",
            "discussion_window",
            "n_agents_total",
            "score_decimals",
            "chapter_batch_size",
        ],
    )

    # scores_long
    ws = wb["scores_long"]
    _format_sheet(ws)
    _apply_number_formats(ws, score_cols=["round", "score"])

    # scores_wide
    ws = wb["scores_wide"]
    _format_sheet(ws)
    # 所有 score_r* 列都设成 0.0
    header = [c.value for c in ws[1]]
    score_cols = [h for h in header if isinstance(h, str) and h.startswith("score_r")]
    _apply_number_formats(ws, score_cols=score_cols)

    # errors
    ws = wb["errors"]
    _format_sheet(ws)

    wb.save(OUT_XLSX)

    print(f"[OK] Found result.json files: {len(result_paths)}")
    print(f"[OK] Exported: {OUT_XLSX}")


if __name__ == "__main__":
    main()